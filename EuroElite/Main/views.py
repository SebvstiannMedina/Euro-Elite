from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.signals import user_logged_in
from django.db import transaction
from django.db.models import Q, Sum, F
from django.db.models.functions import TruncMonth
from django.dispatch import receiver
from django.http import HttpResponseBadRequest, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_POST
from django.contrib.auth import login, logout
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from analytics.utils import track
from .forms import CitaForm, ProductoForm
from .models import Cita, Producto, BloqueHorario
from django.views.decorators.csrf import csrf_exempt
from decimal import Decimal
from django.contrib.admin.views.decorators import staff_member_required

# Local apps
from .forms import CitaForm, DireccionForm, PerfilForm, RegistroForm, EmailLoginForm
from .models import (Carrito,Categoria,ConfigSitio,Direccion,ItemCarrito,ItemPedido,Pago,Pedido,Producto,)
from .decorators import admin_required, mecanico_or_admin_required, repartidor_or_admin_required, asignador_or_admin_required

#temportal -- borrar despues
from django.db import transaction, IntegrityError
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import Cita

from .forms import CitaForm
from .models import Cita, BloqueHorario

def _get_active_cart(user):
    """Devuelve (o crea) el carrito activo del usuario."""
    cart, _ = Carrito.objects.get_or_create(usuario=user, activo=True)
    return cart


def _redirect_back(request, fallback_name):
    """Devuelve un redirect seguro usando next, referer o un fallback."""
    redirect_to = request.POST.get("next") or request.GET.get("next")
    allowed_hosts = {request.get_host()}

    if redirect_to and not url_has_allowed_host_and_scheme(
        url=redirect_to,
        allowed_hosts=allowed_hosts,
        require_https=request.is_secure(),
    ):
        redirect_to = None

    if not redirect_to:
        referer = request.META.get("HTTP_REFERER")
        if referer and url_has_allowed_host_and_scheme(
            url=referer,
            allowed_hosts=allowed_hosts,
            require_https=request.is_secure(),
        ):
            redirect_to = referer

    return redirect(redirect_to or fallback_name)


# ---------- Endpoints de carrito (server) ----------
@login_required
@require_POST
def cart_add(request):
    prod_id = request.POST.get('producto_id')
    try:
        qty = int(request.POST.get('cantidad', 1))
    except (TypeError, ValueError):
        qty = 1

    if not prod_id or qty <= 0:
        return HttpResponseBadRequest("Datos inválidos.")

    producto = get_object_or_404(Producto, id=prod_id)
    cart = _get_active_cart(request.user)

    # Si no tiene stock definido, asumimos 0 para evitar exceso
    stock_disponible = producto.stock if producto.stock is not None else 0
    if stock_disponible <= 0:
        return JsonResponse({"ok": False, "msg": "Producto sin stock disponible."}, status=400)

    # Use the product's discounted price when creating/updating cart items
    item, created = ItemCarrito.objects.get_or_create(
        carrito=cart,
        producto=producto,
        defaults={'cantidad': 0, 'precio_unitario': producto.precio_con_descuento}
    )

    # Calculamos nueva cantidad
    nueva_cantidad = item.cantidad + qty

    # Si supera el stock, lo dejamos en el máximo permitido
    if nueva_cantidad > stock_disponible:
        nueva_cantidad = stock_disponible

    # Si no cambia (ya está al máximo), avisamos al usuario
    if nueva_cantidad == item.cantidad:
        return JsonResponse({
            "ok": False,
            "msg": "Ya has agregado el máximo disponible de este producto."
        }, status=400)

    # Actualizamos cantidad y fijamos el precio unitario al precio (posible) con descuento
    item.cantidad = nueva_cantidad
    item.precio_unitario = producto.precio_con_descuento
    item.save(update_fields=['cantidad', 'precio_unitario'])

    total_items = sum(i.cantidad for i in cart.items.all())
    # Analytics: track add to cart with context
    try:
        track(request, "add_to_cart",
              product_id=producto.id,
              product_name=producto.nombre,
              price=float(producto.precio),
              added_quantity=qty,
              new_quantity=item.cantidad,
              total_items=total_items)
    except Exception:
        pass
    return JsonResponse({
        "ok": True,
        "items": total_items,
        "msg": f"Producto agregado. Cantidad total: {item.cantidad}"
    })


@login_required
@require_POST
def cart_update(request):
    item_id = request.POST.get('item_id')
    try:
        qty = int(request.POST.get('cantidad', 1))
    except (TypeError, ValueError):
        return HttpResponseBadRequest("Cantidad inválida.")

    if not item_id or qty < 0:
        return HttpResponseBadRequest("Datos inválidos.")

    cart = _get_active_cart(request.user)
    item = get_object_or_404(ItemCarrito, id=item_id, carrito=cart)

    if qty == 0:
        item.delete()
    else:
        # Verificar stock disponible
        stock_disponible = item.producto.stock if item.producto.stock is not None else 0
        
        # Si se intenta agregar más del stock disponible, limitar al máximo
        if qty > stock_disponible:
            return JsonResponse({
                "ok": False,
                "msg": f"Solo hay {stock_disponible} unidades disponibles de este producto."
            }, status=400)
        
        item.cantidad = qty
        item.save(update_fields=['cantidad'])

    # Analytics: track cart update
    try:
        cart = _get_active_cart(request.user)
        total_items = sum(i.cantidad for i in cart.items.all())
        total_value = float(sum(i.precio_unitario * i.cantidad for i in cart.items.all()))
        track(request, "update_cart",
              item_id=item.id,
              product_id=item.producto.id,
              new_quantity=qty,
              total_items=total_items,
              total_value=total_value)
    except Exception:
        pass

    return JsonResponse({"ok": True})


@login_required
@require_POST
def cart_remove(request):
    item_id = request.POST.get('item_id')
    if not item_id:
        return HttpResponseBadRequest("Falta item_id.")

    cart = _get_active_cart(request.user)
    item = get_object_or_404(ItemCarrito, id=item_id, carrito=cart)
    item.delete()
    track(request, "remove_from_cart", item_id=item.id, product_id=item.producto.id)  # ← ANALYTICS
    return JsonResponse({"ok": True})


@login_required
def cart_json(request):
    """Devuelve el carrito en JSON para render del front."""
    cart = _get_active_cart(request.user)
    data = []
    for it in cart.items.select_related('producto'):
        data.append({
            "item_id": it.id,
            "producto_id": it.producto.id,
            "nombre": it.producto.nombre,
            "precio": float(it.precio_unitario),
            "cantidad": it.cantidad,
            "subtotal": float(it.precio_unitario * it.cantidad),
        })
    total = sum(d["subtotal"] for d in data)
    total_items = sum(d["cantidad"] for d in data)
    # Analytics: view cart (API)
    try:
        track(request, "view_cart", total=total, count=total_items)
    except Exception:
        pass
    return JsonResponse({"items": data, "total": total, "count": total_items})


# ---------- Checkout ----------
@login_required
@require_POST 
@transaction.atomic
def checkout_crear_pedido_y_pagar(request):
    """
    1) Toma el carrito activo con ítems
    2) Crea Pedido + ItemPedido + Pago(PENDIENTE)
    3) Desactiva el carrito
    4) Redirige a payments:flow_crear_orden con ?pedido_id=...
    """
    from .models import CodigoDescuento
    
    cart = Carrito.objects.filter(usuario=request.user, activo=True).first()
    if not cart or not cart.items.exists():
        messages.error(request, "Tu carrito está vacío.")
        return redirect('carrito_compras')

    addr = Direccion.objects.filter(usuario=request.user).order_by('-predeterminada', '-id').first()

    # Leer método de entrega del formulario (envio | retiro)
    metodo_entrega_form = request.POST.get('metodo_entrega', 'envio').strip().lower()
    
    # Determinar método de entrega para el pedido
    if metodo_entrega_form == 'retiro':
        metodo_entrega = Pedido.MetodoEntrega.RETIRO
    else:
        metodo_entrega = Pedido.MetodoEntrega.DESPACHO

    # Crea Pedido
    pedido = Pedido.objects.create(
        usuario=request.user,
        estado=Pedido.Estado.PENDIENTE,
        metodo_entrega=metodo_entrega,
        metodo_pago=Pedido.MetodoPago.PASARELA,
        direccion_envio=addr,
        direccion_facturacion=addr,
        subtotal=0, descuento=0, envio=0, total=0,
    )

    subtotal = Decimal(0)
    for it in cart.items.select_related('producto'):
        ItemPedido.objects.create(
            pedido=pedido,
            producto=it.producto,
            nombre_producto=it.producto.nombre,
            sku_producto=it.producto.sku,
            precio_unitario=it.precio_unitario,
            cantidad=it.cantidad,
        )
        subtotal += it.precio_unitario * it.cantidad

        # ⚠️ NO descontamos stock aquí - se descuenta solo cuando Flow confirma el pago

    # Calcular costo de envío según método elegido
    if metodo_entrega == Pedido.MetodoEntrega.RETIRO:
        envio_cost = Decimal(0)
    else:
        # Obtener costo de envío configurado (fallback 2990)
        envio_cost = Decimal(2990)
        try:
            cfg = ConfigSitio.objects.first()
            if cfg and getattr(cfg, 'costo_envio_base', None) is not None:
                envio_cost = Decimal(cfg.costo_envio_base)
        except Exception:
            pass

    # Aplicar código de descuento si existe en sesión
    descuento_total = Decimal(0)
    codigo_descuento_str = request.session.get('codigo_descuento', '').strip().upper()
    
    if codigo_descuento_str:
        try:
            codigo_obj = CodigoDescuento.objects.get(codigo=codigo_descuento_str)
            valido, mensaje = codigo_obj.es_valido()
            
            if valido:
                descuento_total = codigo_obj.calcular_descuento(subtotal)
                pedido.codigo_descuento = codigo_obj
                
                # Incrementar usos del código
                codigo_obj.usos_actuales += 1
                codigo_obj.save(update_fields=['usos_actuales'])
                
                # Limpiar sesión
                del request.session['codigo_descuento']
                if 'descuento_aplicado' in request.session:
                    del request.session['descuento_aplicado']
        except CodigoDescuento.DoesNotExist:
            pass

    # Totales
    pedido.subtotal = subtotal
    pedido.envio = envio_cost
    pedido.descuento = descuento_total
    pedido.total = max(Decimal(0), subtotal - descuento_total + envio_cost)
    pedido.save()

    # Pago PENDIENTE
    Pago.objects.create(
        pedido=pedido,
        metodo=Pedido.MetodoPago.PASARELA,
        monto=pedido.total,
        estado=Pago.Estado.PENDIENTE,
    )

    # Cerrar carrito
    cart.activo = False
    cart.save(update_fields=['activo'])
    # Analytics: checkout started (pedido creado)
    try:
        items = []
        for it in pedido.items.all():
            items.append({"product_id": it.producto.id, "cantidad": it.cantidad, "precio_unitario": float(it.precio_unitario)})
        track(request, "checkout_start",
              order_id=pedido.id,
              subtotal=float(pedido.subtotal),
              envio=float(pedido.envio),
              descuento=float(pedido.descuento),
              total=float(pedido.total),
              metodo_entrega=str(pedido.metodo_entrega),
              items=items)
    except Exception:
        pass
    
    # Guardar pedido_id en sesión para recuperarlo después del pago
    request.session['last_order_id'] = pedido.id
    request.session.modified = True
    request.session.save()  # Forzar guardado inmediato
    
    print(f"[CHECKOUT] Pedido {pedido.id} creado y guardado en sesión")
    print(f"[CHECKOUT] Session keys después de guardar: {list(request.session.keys())}")
    
    # Redirigir a Flow
    pay_url = reverse("flow_crear_orden")
    return redirect(f"{pay_url}?pedido_id={pedido.id}")

# ========== PÁGINAS PÚBLICAS ==========
@ensure_csrf_cookie
def home(request):
    ahora = timezone.now()

    # Filtrar productos que tengan stock mayor al mínimo (stock > stock_minimo)
    productos_oferta = (
        Producto.objects.filter(
            activo=True,
            promociones__activa=True,
        )
        .filter(Q(promociones__inicio__isnull=True) | Q(promociones__inicio__lte=ahora))
        .filter(Q(promociones__fin__isnull=True) | Q(promociones__fin__gte=ahora))
        .filter(stock__gt=F('stock_minimo'))  # Solo productos con stock > stock_minimo
        .select_related('categoria')
        .distinct()
        .order_by('-id')[:6]
    )

    productos_normales = (
        Producto.objects.filter(activo=True)
        .filter(stock__gt=F('stock_minimo'))  # Solo productos con stock > stock_minimo
        .exclude(id__in=productos_oferta.values('id'))
        .select_related('categoria')
        .order_by('-id')[:6]
    )

    return render(request, 'taller/main.html', {
        'productos_oferta': productos_oferta,
        'productos_normales': productos_normales,
    })

def producto_detalle(request, pk):
    from .models import Resena
    p = get_object_or_404(
        Producto.objects.select_related('categoria'),
        pk=pk,
        activo=True
    )
    try:
        track(request, "view_product",
              product_id=p.id,
              category=getattr(p.categoria, 'nombre', None),
              price=float(p.precio) if getattr(p, 'precio', None) is not None else None)
    except Exception:
        pass
    
    # Obtener reseñas aprobadas
    resenas = Resena.objects.filter(producto=p, aprobada=True).select_related('usuario').order_by('-creado')
    
    # Verificar si el usuario puede reseñar (ha comprado el producto)
    puede_resenar = False
    if request.user.is_authenticated:
        puede_resenar = (
            ItemPedido.objects.filter(
                pedido__usuario=request.user,
                pedido__estado='ENTREGADO',
                producto=p
            ).exists() and 
            not Resena.objects.filter(usuario=request.user, producto=p).exists()
        )
    
    return render(request, 'taller/producto_detalle.html', {
        'producto': p,
        'p': p,
        'resenas': resenas,
        'puede_resenar': puede_resenar
    })


def nosotros(request):
    from .models import FotoNosotros, Testimonio
    fotos = FotoNosotros.objects.filter(activa=True).order_by('orden', '-creado')
    resenas = Testimonio.objects.filter(aprobada=True).order_by('-creado')[:6]
    return render(request, 'taller/nosotros.html', {'fotos': fotos, 'resenas': resenas})

def equipo(request):
    return render(request, 'taller/equipo.html')

def productos(request):
    # Solo mostrar productos con stock > stock_minimo
    productos = Producto.objects.filter(activo=True).filter(stock__gt=F('stock_minimo')).select_related("categoria")
    categorias = Categoria.objects.filter(activa=True)

    for p in productos:
        p.promocion = p.promocion_vigente
        p.precio_descuento = p.precio_con_descuento

    try:
        track(request, "list_products", count=productos.count())
    except Exception:
        pass

    return render(request, "taller/productos.html", {
        "productos": productos,
        "categorias": categorias,
    })


# ========== LOGIN ==========
from django.contrib.auth.views import LoginView, LogoutView
from django.contrib import messages
from .forms import EmailAuthenticationForm
from .forms import RegistroForm
from django.shortcuts import render, redirect

class CustomLoginView(LoginView):
    template_name = "taller/login.html"
    authentication_form = EmailAuthenticationForm

    def form_valid(self, form):
        user = form.get_user()

        # Si el usuario está bloqueado, detener login
        if getattr(user, 'bloqueado', False):
            messages.error(
                self.request,
                "⚠️ Tu cuenta ha sido bloqueada por un administrador. "
                "Si crees que se trata de un error, contacta con el soporte de Euro Elite."
            )
            return redirect('login')

        # Login normal
        messages.success(self.request, f"Bienvenido: {user.first_name or user.email}")
        track(self.request, "login", user_id=user.id)
        return super().form_valid(form)

    def form_invalid(self, form):
        """
        Este método se ejecuta cuando el formulario NO es válido,
        es decir, cuando las credenciales son incorrectas o el backend
        rechazó al usuario bloqueado.
        """
        email = self.request.POST.get('username') or self.request.POST.get('email')

        from Main.models import Usuario  
        try:
            usuario = Usuario.objects.get(email=email)
            if usuario.bloqueado:
                messages.error(
                    self.request,
                    "⚠️ Tu cuenta ha sido bloqueada por un administrador. "
                    "No puedes iniciar sesión hasta que sea desbloqueada."
                )
                return redirect('login')
        except Usuario.DoesNotExist:
            pass

        # Si no está bloqueado, mostrar error estándar
        messages.error(self.request, "Correo o contraseña incorrectos.")
        return super().form_invalid(form)
# ============ LOGOUT ============
class CustomLogoutView(LogoutView):
    next_page = "login"
# ========== REGISTRO ==========

def registro(request):
    if request.user.is_authenticated:
        return redirect('home')  

    if request.method == 'POST':
        form = RegistroForm(request.POST)
        if form.is_valid():
            user = form.save()
            try:
                track(request, "user_registered", user_id=getattr(user, 'id', None), email=getattr(user, 'email', None))
            except Exception:
                pass
            messages.success(request, 'Registro exitoso. Inicia sesión para continuar.')
            return redirect('login')
        else:
            
            print("Errores del formulario:", form.errors)
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = RegistroForm()

    return render(request, 'taller/registro.html', {'form': form})

# ========== PERFIL ==========

@login_required
def perfil(request):
    usuario = request.user
    addr = Direccion.objects.filter(usuario=request.user, tipo="ENVIO")\
        .order_by('-predeterminada', '-id').first()

# Si no existe dirección en DB, crear una con los datos del checkout
    metodo_entrega_form = request.POST.get('metodo_entrega', 'envio').strip().lower()
    metodo_entrega = Pedido.MetodoEntrega.RETIRO if metodo_entrega_form == 'retiro' else Pedido.MetodoEntrega.DESPACHO
    
    if not addr and metodo_entrega == Pedido.MetodoEntrega.DESPACHO:
        addr = Direccion.objects.create(
            usuario=request.user,
            tipo="ENVIO",
            nombre_completo=f"{request.user.first_name} {request.user.last_name}",
            telefono=request.POST.get("telefono_envio", ""),
            linea1=request.POST.get("direccion_envio", ""),
            linea2=request.POST.get("direccion_envio2", ""),
            comuna=request.POST.get("comuna_envio", ""),
            ciudad=request.POST.get("ciudad_envio", "Santiago"),
            region=request.POST.get("region_envio", ""),
            predeterminada=True
        )

    if request.method == "POST":
        perfil_form = PerfilForm(request.POST, instance=usuario)
        direccion_form = DireccionForm(request.POST, instance=addr)

        if perfil_form.is_valid() and direccion_form.is_valid():
            # Guardar perfil del usuario
            perfil_form.save()

            # Guardar dirección (crear o actualizar)
            direccion = direccion_form.save(commit=False)
            direccion.usuario = usuario
            
            # Rellenar campos requeridos de Direccion desde Usuario
            direccion.nombre_completo = f"{usuario.first_name} {usuario.last_name}".strip()
            direccion.telefono = usuario.telefono or ""
            
            # Si no hay dirección existente, marcar como predeterminada
            if not addr:
                direccion.predeterminada = True
            
            direccion.save()
            
            try:
                track(request, "profile_updated", user_id=request.user.id)
            except Exception:
                pass
            messages.success(request, "Perfil actualizado correctamente ✔")
            return redirect("perfil")
    else:
        perfil_form = PerfilForm(instance=usuario)
        
        # Si no existe dirección, crear un formulario vacío
        if addr:
            direccion_form = DireccionForm(instance=addr)
        else:
            direccion_form = DireccionForm()

    return render(request, "taller/perfil.html", {
        "form": perfil_form,
        "direccion_form": direccion_form,
        "user": usuario,
        "addr": addr,
    })

# ========== LOGOUT ==========
@login_required
def logout_view(request):
    auth_logout(request)
    return redirect('home')


#Vistas de agenda
@login_required
def agendar(request):
    if request.method == "POST":
        form = CitaForm(request.POST, user=request.user)

        if form.is_valid():
            try:
                with transaction.atomic():

                    bloque_id = form.cleaned_data["bloque"].id
                    bloque = BloqueHorario.objects.select_for_update().get(id=bloque_id)

                    if bloque.bloqueado or Cita.objects.filter(bloque=bloque).exists():
                        form.add_error("bloque", "Este bloque ya está reservado.")
                        raise ValueError("Bloque reservado")

                    cita = Cita(
                        usuario=request.user,
                        servicio=form.cleaned_data["servicio"],
                        bloque=bloque,
                        estado=Cita.Estado.RESERVADA,
                        a_domicilio=form.cleaned_data.get("a_domicilio", False),
                        direccion_domicilio=form.cleaned_data.get("direccion_domicilio", "")
                    )
                    cita.save()

                    bloque.bloqueado = True
                    bloque.save(update_fields=["bloqueado"])

            except ValueError:
                pass

            except IntegrityError:
                form.add_error("bloque", "Este bloque ya está reservado.")

            else:
                try:
                    track(request, "appointment_booked", servicio=form.cleaned_data.get("servicio"), usuario_id=request.user.id, bloque_id=bloque.id, inicio=str(bloque.inicio))
                except Exception:
                    pass
                messages.success(request, "Tu cita fue reservada correctamente ✅")
                return redirect("mis_citas")

    else:
        form = CitaForm(user=request.user)

    return render(request, "taller/agendar.html", {"form": form})


@login_required
def mis_citas(request):
    citas = (
        Cita.objects
        .filter(usuario=request.user)
        .select_related("servicio", "bloque")
        .order_by("bloque__inicio")
    )
    return render(request, "taller/mis_citas.html", {"citas": citas})


@login_required
def anular_cita(request, cita_id):
    from .models import HoraDisponible
    import json
    
    base_queryset = Cita.objects.select_related("bloque")
    if request.user.is_staff:
        cita = get_object_or_404(base_queryset, id=cita_id)
    else:
        cita = get_object_or_404(base_queryset, id=cita_id, usuario=request.user)

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/json'

    if cita.estado == Cita.Estado.RESERVADA:
        cita.estado = Cita.Estado.CANCELADA
        if cita.bloque:
            # Restaurar la hora disponible
            fecha = cita.bloque.inicio.date()
            hora = cita.bloque.inicio.time()
            HoraDisponible.objects.filter(fecha=fecha, hora=hora).update(disponible=True)
            
            # Desbloquear el bloque
            cita.bloque.bloqueado = False
            cita.bloque.save(update_fields=["bloqueado"])
        cita.save(update_fields=["estado"])
        try:
            track(request, "appointment_canceled", cita_id=cita.id, usuario_id=(request.user.id if request.user.is_authenticated else None))
        except Exception:
            pass
        
        if is_ajax:
            return JsonResponse({"success": True, "message": "La cita fue cancelada correctamente."})
        messages.success(request, "La cita fue cancelada correctamente.")
    else:
        if is_ajax:
            return JsonResponse({"success": False, "error": "Solo puedes cancelar citas que estén reservadas."})
        messages.warning(request, "Solo puedes cancelar citas que estén reservadas.")

    if is_ajax:
        return JsonResponse({"success": True})
    return _redirect_back(request, "mis_citas")


@staff_member_required
@require_POST
def avanzar_estado_cita(request, cita_id):
    cita = get_object_or_404(Cita, id=cita_id)

    transiciones = {
        Cita.Estado.RESERVADA: (Cita.Estado.EN_PROCESO, "La cita ahora está en proceso."),
        Cita.Estado.EN_PROCESO: (Cita.Estado.COMPLETADA, "La cita fue completada."),
    }

    siguiente = transiciones.get(cita.estado)

    if not siguiente:
        messages.warning(request, "Esta cita no puede avanzar de estado.")
        return _redirect_back(request, "admin_agendamiento")

    nuevo_estado, mensaje = siguiente
    cita.estado = nuevo_estado
    cita.save(update_fields=["estado"])
    try:
        track(request, "appointment_state_changed", cita_id=cita.id, nuevo_estado=str(nuevo_estado), actor_id=(request.user.id if request.user.is_authenticated else None))
    except Exception:
        pass
    messages.success(request, mensaje)

    return _redirect_back(request, "admin_agendamiento")

def nueva_contrasena(request):
    return render(request, 'taller/nueva_contrasena.html')

def pago(request):
    return render(request, 'taller/pago.html')

def carrito_compras(request):
    return render(request, 'taller/carrito_compras.html')

def terminos(request):
    return render(request, 'taller/terminos.html')

def privacidad(request):
    return render(request, 'taller/privacidad.html')

from .forms import ProductoForm
from .models import Producto
from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
@login_required
@require_POST
def crear_promocion(request):
    """Crea una nueva promoción via AJAX"""
    try:
        from .models import Promocion
        from decimal import Decimal
        
        nombre = request.POST.get('nombre', '').strip()
        tipo = request.POST.get('tipo', 'PORCENTAJE')
        valor = request.POST.get('valor', '0')
        inicio = request.POST.get('inicio', None)
        fin = request.POST.get('fin', None)
        activa = request.POST.get('activa') == 'on'
        
        if not nombre:
            return JsonResponse({'success': False, 'error': 'El nombre es obligatorio'})
        
        if tipo not in ['PORCENTAJE', 'MONTO']:
            return JsonResponse({'success': False, 'error': 'Tipo de promoción inválido'})
        
        try:
            valor_decimal = Decimal(valor)
            if valor_decimal < 0:
                return JsonResponse({'success': False, 'error': 'El valor debe ser positivo'})
            if tipo == 'PORCENTAJE' and valor_decimal > 100:
                return JsonResponse({'success': False, 'error': 'El porcentaje no puede ser mayor a 100'})
        except:
            return JsonResponse({'success': False, 'error': 'Valor inválido'})
        
        # Convertir fechas
        inicio_dt = None
        fin_dt = None
        if inicio:
            try:
                from django.utils.dateparse import parse_datetime
                inicio_dt = parse_datetime(inicio)
            except:
                pass
        
        if fin:
            try:
                from django.utils.dateparse import parse_datetime
                fin_dt = parse_datetime(fin)
            except:
                pass
        
        # Crear promoción
        promocion = Promocion.objects.create(
            nombre=nombre,
            tipo=tipo,
            valor=valor_decimal,
            inicio=inicio_dt,
            fin=fin_dt,
            activa=activa
        )
        
        return JsonResponse({
            'success': True,
            'promocion': {
                'id': promocion.id,
                'nombre': promocion.nombre,
                'tipo': promocion.tipo,
                'valor': str(promocion.valor)
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@staff_member_required
@login_required
def agregar_editar(request, pk=None): 
    if pk:  # Editar producto
        producto = get_object_or_404(Producto, pk=pk)
    else:   # Agregar producto
        producto = None

    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            producto = form.save()
            messages.success(request, 'Producto guardado correctamente.')
            # 
            if pk:
                return redirect('editar_producto', pk=producto.pk)
           
            return redirect('agregar_editar')
    else:
        form = ProductoForm(instance=producto)

    productos = Producto.objects.all().order_by('-id')
    return render(request, 'taller/agregar_editar.html', {
        'form': form,
        'productos': productos,
        'editando': producto is not None
    })


def prueba(request):
    return render(request, 'taller/prueba.html')

@csrf_exempt
def compra_exitosa(request, pedido_id=None):
    """
    Muestra la página de compra exitosa con los detalles del pedido.
    PRIORIDAD: Encontrar el pedido primero, autenticación después.
    """
    print(f"[COMPRA_EXITOSA] ========== INICIO ==========")
    print(f"[COMPRA_EXITOSA] pedido_id recibido: {pedido_id}")
    print(f"[COMPRA_EXITOSA] Usuario autenticado: {request.user.is_authenticated}")
    print(f"[COMPRA_EXITOSA] User: {request.user}")
    print(f"[COMPRA_EXITOSA] Session keys: {list(request.session.keys())}")
    print(f"[COMPRA_EXITOSA] GET params: {dict(request.GET)}")
    
    pedido = None
    pago = None
    
    # PASO 1: Intentar obtener pedido_id de TODAS las fuentes posibles
    if not pedido_id:
        # Desde la sesión
        pedido_id = request.session.get('last_order_id')
        print(f"[COMPRA_EXITOSA] pedido_id desde sesión: {pedido_id}")
    
    # PASO 2: Intentar cargar el pedido (con o sin autenticación)
    if pedido_id:
        try:
            # Intentar cargar el pedido SIN filtrar por usuario
            pedido = Pedido.objects.select_related('usuario').prefetch_related('items__producto').get(id=pedido_id)
            pago = getattr(pedido, 'pago', None)
            print(f"[COMPRA_EXITOSA] ✅ Pedido {pedido.id} encontrado para usuario {pedido.usuario.email}")
            print(f"[COMPRA_EXITOSA] Estado del pedido: {pedido.estado}")
            
            if pedido.estado == Pedido.Estado.CANCELADO:
                print(f"[COMPRA_EXITOSA] 🚫 Pedido CANCELADO - redirigiendo a compra_rechazada")
                return redirect('compra_rechazada', pedido_id=pedido.id)
            
            # Guardar en sesión para futuras visitas
            request.session['last_order_id'] = pedido.id
            
        except Pedido.DoesNotExist:
            print(f"[COMPRA_EXITOSA] ❌ Pedido {pedido_id} NO existe en BD")
            pedido = None
    
    # PASO 3: Si NO hay pedido y el usuario está autenticado, buscar el más reciente
    if not pedido and request.user.is_authenticated:
        print(f"[COMPRA_EXITOSA] Buscando pedido más reciente para usuario autenticado")
        pedido = Pedido.objects.filter(
            usuario=request.user
        ).select_related('usuario').prefetch_related('items__producto').order_by('-creado').first()
        
        if pedido:
            pago = getattr(pedido, 'pago', None)
            print(f"[COMPRA_EXITOSA] ✅ Pedido más reciente encontrado: {pedido.id}")
    
    # PASO 4: Si AÚN no hay pedido, mostrar mensaje
    if not pedido:
        print(f"[COMPRA_EXITOSA] ⚠️ No se encontró ningún pedido")
        messages.info(request, "Tu pago está siendo procesado. Recibirás una confirmación por correo.")
    
    # PASO 5: Preparar contexto (funciona con o sin pedido)
    comprobante_url = pago.comprobante_url if pago and hasattr(pago, 'comprobante_url') else None
    numero_pedido = getattr(pedido, 'id', None) if pedido else None
    total = getattr(pedido, 'total', None) if pedido else None
    metodo_pago = getattr(pago, 'metodo', None) if pago else None

    # Construir items con precios finales (prorrateo de descuento por ítem solo para visualización)
    items_con_precios = None
    if pedido:
        try:
            from decimal import Decimal, ROUND_HALF_UP
            items = list(pedido.items.all())
            sub_total_pedido = sum((i.subtotal for i in items), Decimal(0))
            descuento_total = getattr(pedido, 'descuento', Decimal(0)) or Decimal(0)

            if items and sub_total_pedido > 0 and descuento_total > 0:
                items_con_precios = []
                descuento_restante = descuento_total
                # Prorratear el descuento con redondeo a pesos; ajustar en el último ítem para cuadrar
                for idx, it in enumerate(items):
                    sub_it = it.subtotal
                    if idx < len(items) - 1:
                        propor = (sub_it / sub_total_pedido)
                        desc_it = (descuento_total * propor).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
                        # No exceder el subtotal del ítem
                        if desc_it > sub_it:
                            desc_it = sub_it
                        descuento_restante -= desc_it
                    else:
                        # Al último ítem se le asigna el descuento restante para cuadrar
                        desc_it = max(Decimal(0), min(descuento_restante, sub_it))
                    sub_final = sub_it - desc_it
                    unit_final = (sub_final / it.cantidad) if it.cantidad else Decimal(0)
                    items_con_precios.append({
                        'nombre': it.nombre_producto,
                        'cantidad': it.cantidad,
                        'precio_unitario_final': unit_final,
                        'precio_unitario_original': it.precio_unitario,
                        'subtotal_final': sub_final,
                        'subtotal_original': sub_it,
                        'ahorro': (sub_it - sub_final),
                    })
        except Exception:
            items_con_precios = None
    
    # Usuario para el template (puede ser el del pedido o el autenticado)
    template_user = request.user if request.user.is_authenticated else (pedido.usuario if pedido else None)
    
    print(f"[COMPRA_EXITOSA] Renderizando template con pedido_id={numero_pedido}")
    print(f"[COMPRA_EXITOSA] ========== FIN ==========")
    
    # Analytics: purchase completed (if pedido available)
    try:
        if pedido:
            items = []
            for it in pedido.items.all():
                items.append({"product_id": it.producto.id, "cantidad": it.cantidad, "subtotal": float(it.subtotal)})
            track(request, "purchase_completed",
                  order_id=pedido.id,
                  total=float(pedido.total) if pedido.total is not None else None,
                  metodo_pago=str(metodo_pago) if metodo_pago else None,
                  items=items)
    except Exception:
        pass

    return render(request, 'taller/compra_exitosa.html', {
        'user': template_user,
        'pedido': pedido,
        'numero_pedido': numero_pedido,
        'total': total,
        'metodo_pago': metodo_pago,
        'comprobante_url': comprobante_url,
        'items_con_precios': items_con_precios,
    })


@csrf_exempt
def compra_rechazada(request, pedido_id=None):
    """
    Muestra la página de compra rechazada cuando el pago falla.
    """
    print(f"[COMPRA_RECHAZADA] ========== INICIO ==========")
    print(f"[COMPRA_RECHAZADA] pedido_id recibido: {pedido_id}")
    print(f"[COMPRA_RECHAZADA] Usuario autenticado: {request.user.is_authenticated}")
    
    pedido = None
    pago = None
    motivo_rechazo = None
    
    # Intentar obtener pedido_id de sesión si no se proporciona
    if not pedido_id:
        pedido_id = request.session.get('last_order_id')
        print(f"[COMPRA_RECHAZADA] pedido_id desde sesión: {pedido_id}")
    
    # Cargar el pedido
    if pedido_id:
        try:
            pedido = Pedido.objects.select_related('usuario').prefetch_related('items__producto').get(id=pedido_id)
            pago = getattr(pedido, 'pago', None)
            print(f"[COMPRA_RECHAZADA] ✅ Pedido {pedido.id} encontrado, estado: {pedido.estado}")
            
            # Intentar obtener el motivo del rechazo desde la respuesta de Flow
            if pago and hasattr(pago, 'flow_response') and pago.flow_response:
                try:
                    import json
                    flow_data = json.loads(pago.flow_response.replace("'", '"'))
                    last_error = flow_data.get('lastError', {})
                    if last_error and last_error.get('message'):
                        motivo_rechazo = last_error.get('message')
                except:
                    pass
                    
        except Pedido.DoesNotExist:
            print(f"[COMPRA_RECHAZADA] ❌ Pedido {pedido_id} NO existe")
    
    # Si no hay pedido y el usuario está autenticado, buscar el más reciente cancelado
    if not pedido and request.user.is_authenticated:
        pedido = Pedido.objects.filter(
            usuario=request.user,
            estado=Pedido.Estado.CANCELADO
        ).order_by('-creado').first()
        
        if pedido:
            pago = getattr(pedido, 'pago', None)
            print(f"[COMPRA_RECHAZADA] ✅ Pedido cancelado más reciente: {pedido.id}")
    
    numero_pedido = getattr(pedido, 'id', None) if pedido else None
    total = getattr(pedido, 'total', None) if pedido else None
    template_user = request.user if request.user.is_authenticated else (pedido.usuario if pedido else None)
    
    print(f"[COMPRA_RECHAZADA] Renderizando template con pedido_id={numero_pedido}")
    print(f"[COMPRA_RECHAZADA] ========== FIN ==========")
    # Analytics: purchase failed
    try:
        track(request, "purchase_failed", order_id=numero_pedido, total=float(total) if total else None, reason=(motivo_rechazo or None))
    except Exception:
        pass

    return render(request, 'taller/compra_rechazada.html', {
        'user': template_user,
        'pedido': pedido,
        'numero_pedido': numero_pedido,
        'total': total,
        'motivo_rechazo': motivo_rechazo or 'El pago fue rechazado por el medio de pago.',
    })



def ofertas(request):
    return render(request, 'taller/ofertas.html')

def retiro_despacho(request):
    return render(request, 'taller/retiro_despacho.html')

from django.contrib.admin.views.decorators import staff_member_required


@admin_required
def admin_agendamientos(request):
    citas = (
        Cita.objects
        .select_related('usuario', 'servicio', 'bloque')
        .order_by('-bloque__inicio')
    )
    agendamientos_activos = citas.filter(
        estado__in=[Cita.Estado.RESERVADA, Cita.Estado.EN_PROCESO]
    ).count()

    return render(
        request,
        'taller/admin_agendamientos.html',
        {
            'citas': citas,
            'agendamientos_activos': agendamientos_activos,
        },
    )

@admin_required
def admin_configuracion(request):
    return render(request, 'taller/admin_configuracion.html')

@asignador_or_admin_required
def admin_asignacion(request):
    """
    Vista para que asignadores puedan asignar pedidos a repartidores.
    """
    # Estados válidos para asignación
    estados_asignables = [
        Pedido.Estado.PAGADO,
        Pedido.Estado.PREPARACION,
    ]
    
    # Obtener pedidos sin asignar o que necesitan ser reasignados
    pedidos_sin_asignar = Pedido.objects.filter(
        estado__in=estados_asignables,
        asignado_a__isnull=True
    ).select_related(
        'usuario',
        'direccion_envio',
        'pago'
    ).prefetch_related(
        'items__producto'
    ).order_by('creado')
    
    # Obtener todos los repartidores activos
    repartidores = Usuario.objects.filter(
        rol='REPARTIDOR',
        bloqueado=False
    ).order_by('first_name', 'last_name')
    
    # Manejar asignación POST
    if request.method == 'POST' and request.user.rol in ['ADMIN', 'ASIGNADOR']:
        pedido_id = request.POST.get('pedido_id')
        repartidor_id = request.POST.get('repartidor_id')
        
        try:
            pedido = Pedido.objects.get(id=pedido_id)
            repartidor = Usuario.objects.get(id=repartidor_id, rol='REPARTIDOR')
            
            # Asignar pedido
            pedido.asignado_a = repartidor
            pedido.asignado_por = request.user
            pedido.save()
            
            messages.success(
                request, 
                f"✅ Pedido #{pedido.id} asignado a {repartidor.get_full_name()}."
            )
        except (Pedido.DoesNotExist, Usuario.DoesNotExist):
            messages.error(request, "❌ Error: Pedido o repartidor no encontrado.")
        
        return redirect('admin_asignacion')
    
    context = {
        'pedidos_sin_asignar': pedidos_sin_asignar,
        'repartidores': repartidores,
    }
    
    return render(request, 'taller/admin_asignacion.html', context)

@admin_required
def admin_pedidos(request):
    # Obtener solo pedidos pagados correctamente (excluyendo pendientes y cancelados)
    estados_validos = [
        Pedido.Estado.PAGADO,
        Pedido.Estado.PREPARACION,
        Pedido.Estado.EN_RUTA,
        Pedido.Estado.ENVIADO,
        Pedido.Estado.ENTREGADO,
    ]
    
    pedidos = Pedido.objects.filter(
        estado__in=estados_validos
    ).select_related(
        'usuario', 
        'direccion_envio',
        'direccion_facturacion',
        'pago'
    ).prefetch_related(
        'items__producto'
    ).order_by('-creado')
    
    # Agrupar pedidos por usuario para la vista
    from itertools import groupby
    pedidos_por_usuario = []
    
    # Agrupar por usuario.email
    for email, pedidos_grupo in groupby(pedidos, key=lambda p: p.usuario.email if p.usuario else 'Sin usuario'):
        pedidos_lista = list(pedidos_grupo)
        pedidos_por_usuario.append({
            'grouper': email,
            'list': pedidos_lista
        })
    
    return render(request, 'taller/admin_pedidos.html', {
        'pedidos': pedidos,
        'pedidos_por_usuario': pedidos_por_usuario,
    })

def asignar_pedidos(request):
    return render(request, 'taller/asignar_pedidos.html')

from django.contrib.auth import get_user_model
from django.contrib import messages
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.admin.views.decorators import staff_member_required

Usuario = get_user_model()  # tu modelo custom de usuario

from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import render
from .models import Usuario

@admin_required
def admin_usuarios(request):
    # Obtener parámetros de búsqueda
    q = request.GET.get('q', '')
    rol = request.GET.get('rol', '')

    # Procesar cambio de rol (POST)
    if request.method == 'POST' and request.POST.get('action') == 'change_role':
        usuario_id = request.POST.get('usuario_id')
        nuevo_rol = request.POST.get('nuevo_rol')
        try:
            usuario_obj = Usuario.objects.get(id=usuario_id)
        except Usuario.DoesNotExist:
            messages.error(request, "Usuario no encontrado.")
            return redirect('admin_usuarios')

        # Prevenciones de seguridad
        if usuario_obj.is_superuser:
            messages.error(request, "No puedes cambiar el rol de un superusuario.")
            return redirect('admin_usuarios')

        if usuario_obj == request.user:
            messages.error(request, "No puedes cambiar tu propio rol desde aquí.")
            return redirect('admin_usuarios')

        # Validar rol
        roles_validos = [choice[0] for choice in Usuario.Rol.choices]
        if nuevo_rol not in roles_validos:
            messages.error(request, "Rol inválido seleccionado.")
            return redirect('admin_usuarios')

        usuario_obj.rol = nuevo_rol
        # No modificamos is_superuser/is_staff aquí; solo el campo rol
        usuario_obj.save()
        messages.success(request, f"Rol de {usuario_obj.get_full_name() or usuario_obj.email} actualizado a {nuevo_rol}.")
        return redirect('admin_usuarios')

    # Base queryset (todos los usuarios)
    usuarios = Usuario.objects.all().order_by('-date_joined')

    # --- Filtro de búsqueda ---
    if q:
        usuarios = usuarios.filter(
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(email__icontains=q) |
            Q(username__icontains=q)
        )

    # --- Filtro por rol ---
    if rol:
        usuarios = usuarios.filter(rol=rol)

    # --- Paginación (20 por página) ---
    paginator = Paginator(usuarios, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        "usuarios": page_obj,  # importante: para que el HTML funcione igual
        "page_obj": page_obj,
        "q": q,
        "rol": rol,
    }

    return render(request, 'taller/admin_usuarios.html', context)



# Detalle de un usuario
@staff_member_required
def detalle_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    return render(request, 'taller/detalle_usuario.html', {"usuario": usuario})

# Eliminar un usuario
@staff_member_required
def eliminar_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    if usuario.is_superuser:
        messages.error(request, "No puedes eliminar a un superusuario.")
    else:
        usuario.delete()
        messages.success(request, "Usuario eliminado correctamente.")
    return redirect("admin_usuarios")



def mis_pedidos(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    # Mostrar solo pedidos realmente comprados o en proceso de despacho
    estados_visibles = [
        Pedido.Estado.PAGADO,
        Pedido.Estado.PREPARACION,
        Pedido.Estado.EN_RUTA if hasattr(Pedido.Estado, 'EN_RUTA') else 'EN_RUTA',
        Pedido.Estado.ENVIADO if hasattr(Pedido.Estado, 'ENVIADO') else 'ENVIADO',
        Pedido.Estado.ENTREGADO,
    ]

    pedidos_qs = (
        request.user.pedidos
        .filter(estado__in=estados_visibles)
        .select_related('pago')
        .prefetch_related('items__producto__imagenes')
        .order_by('-creado')
    )
    
    pedidos = []
    for p in pedidos_qs:
        # Obtener el primer item para mostrar la imagen principal
        primer_item = p.items.first()
        imagen_url = ''
        titulo = 'Detalle de compra'
        descripcion = ''
        cantidad_total = sum(item.cantidad for item in p.items.all())
        num_items_diferentes = p.items.count()
        
        if primer_item:
            # Imagen del primer producto - primero intentar la imagen directa del producto
            if primer_item.producto.imagen:
                imagen_url = primer_item.producto.imagen.url
            else:
                # Si no tiene imagen directa, buscar en las imágenes relacionadas
                primera_imagen = primer_item.producto.imagenes.first()
                if primera_imagen:
                    imagen_url = primera_imagen.imagen.url
            
            # Título y descripción
            if num_items_diferentes == 1:
                titulo = primer_item.nombre_producto
                descripcion = f"SKU: {primer_item.sku_producto}"
            else:
                titulo = f"{primer_item.nombre_producto} y {num_items_diferentes - 1} producto(s) más"
                descripcion = f"Pedido con {num_items_diferentes} productos"
        
        # Calcular fecha de entrega estimada (7 días después de creado si es despacho)
        fecha_entrega = None
        if p.metodo_entrega == Pedido.MetodoEntrega.DESPACHO and p.estado in [Pedido.Estado.PAGADO, Pedido.Estado.PREPARACION, Pedido.Estado.ENVIADO]:
            fecha_entrega = p.creado + timezone.timedelta(days=7)
        
        pedidos.append({
            'id': p.id,
            'fecha': p.creado,
            'fecha_entrega': fecha_entrega,
            'estado': p.get_estado_display(),
            'imagen_url': imagen_url,
            'titulo': titulo,
            'descripcion': descripcion,
            'cantidad': cantidad_total,
            'vendedor': 'Euro Elite',
            'detalle_url': reverse('compra_exitosa_detalle', args=[p.id]),
            'recomprar_url': None,  # Se puede implementar más adelante
        })
    
    return render(request, 'taller/mis_pedidos.html', {
        'pedidos': pedidos,
        'DEBUG': settings.DEBUG,
    })


def resumen_compra(request):
    shipping_cost = 2990
    try:
        cfg = ConfigSitio.objects.first()
        if cfg and cfg.costo_envio_base is not None:
            shipping_cost = int(cfg.costo_envio_base)
    except Exception:
        pass
    return render(request, 'taller/resumen_compra.html', {
        'shipping_cost': shipping_cost,
    })


from django.contrib.auth.decorators import login_required

@login_required
def confirmacion_datos(request):
    user = request.user
    rut = request.session.get('checkout_rut', '')
    addr = Direccion.objects.filter(usuario=user).order_by('-predeterminada', '-id').first()

    if request.method == 'POST':
        rut = request.POST.get('rut', '').strip()
        nombre = request.POST.get('nombre', '').strip()
        apellido = request.POST.get('apellido', '').strip()
        email = request.POST.get('email', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        direccion_txt = request.POST.get('direccion', '').strip()
        direccion2_txt = request.POST.get('direccion2', '').strip()
        region = request.POST.get('region', '').strip()
        comuna = request.POST.get('comuna', '').strip()
        notas = request.POST.get('notas', '').strip()

        if nombre:
            user.first_name = nombre[:25] 
        if apellido:
            user.last_name = apellido[:25]
        if email:
            email_clean = email.replace(' ', '').lower()
            for char in ['<', '>', '"', "'", ';', '\r', '\n', '\t']:
                email_clean = email_clean.replace(char, '')
            if len(email_clean) <= 120:
                user.email = email_clean
        if telefono:
            telefono_digits = ''.join(filter(str.isdigit, telefono))
            if len(telefono_digits) == 8:
                user.telefono = telefono_digits
        if rut:
            user.rut = rut
        
        user.save(update_fields=['first_name', 'last_name', 'email', 'telefono', 'rut'])

        if not addr:
            addr = Direccion(usuario=user, tipo=Direccion.Tipo.ENVIO)
        
        if direccion_txt:
            addr.linea1 = direccion_txt[:65] 
        if direccion2_txt:
            addr.linea2 = direccion2_txt[:65]
        if telefono:
            telefono_digits = ''.join(filter(str.isdigit, telefono))
            if len(telefono_digits) == 8:
                addr.telefono = telefono_digits
        if region:
            addr.region = region
        if comuna:
            addr.comuna = comuna
        if not addr.ciudad:
            addr.ciudad = 'Santiago'
        
        addr.predeterminada = True
        addr.save()

        request.session['checkout_rut'] = rut

        return redirect('pago')

    try:
        cart = _get_active_cart(user)
    except Exception:
        cart = None

    subtotal_dec = Decimal(0)
    if cart:
        try:
            for it in cart.items.all():
                price = it.precio_unitario or Decimal(0)
                qty = int(getattr(it, 'cantidad', 0) or 0)
                subtotal_dec += (price * qty)
        except Exception:
            pass

    shipping_cost_int = 2990
    try:
        cfg = ConfigSitio.objects.first()
        if cfg and getattr(cfg, 'costo_envio_base', None) is not None:
            shipping_cost_int = int(cfg.costo_envio_base)
    except Exception:
        pass

    envio_dec = Decimal(shipping_cost_int)

    # Leer descuento y código de la sesión
    descuento_aplicado = Decimal(request.session.get('descuento_aplicado', '0'))
    codigo_descuento = request.session.get('codigo_descuento', '')

    total_dec = (subtotal_dec or Decimal(0)) - descuento_aplicado + envio_dec
    if total_dec < 0:
        total_dec = Decimal(0)

    def _format_clp(amount: Decimal) -> str:
        try:
            return "$" + f"{int(amount):,}".replace(",", ".")
        except Exception:
            return "-"

    context = {
        'rut': rut,
        'addr': addr,
        'user': user,
        'subtotal': _format_clp(subtotal_dec),
        'envio': _format_clp(envio_dec),
        'descuento': _format_clp(descuento_aplicado) if descuento_aplicado else None,
        'codigo_descuento': codigo_descuento,
        'total': _format_clp(total_dec),
    }

    return render(request, 'taller/confirmacion_datos.html', context)

def olvide_contra(request):
    return render(request, 'taller/olvide_contra.html')

from datetime import datetime
from decimal import Decimal
from django.shortcuts import render
from django.db.models import Sum, F, ExpressionWrapper, DecimalField
from django.db.models.functions import TruncMonth

from Main.models import Pedido, ItemPedido


def estadisticas(request):
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    estado = request.GET.get("estado")

    pedidos = Pedido.objects.all().order_by("-creado")


    if fecha_inicio:
        pedidos = pedidos.filter(creado__date__gte=fecha_inicio)
    if fecha_fin:
        pedidos = pedidos.filter(creado__date__lte=fecha_fin)
    if estado:
        pedidos = pedidos.filter(estado=estado)

  
    total_ganancia = pedidos.aggregate(total=Sum("ganancia"))["total"] or 0
    total_iva = pedidos.aggregate(total=Sum("iva"))["total"] or 0
    cantidad_pedidos = pedidos.count()

    ingresos_por_mes = (
        pedidos.annotate(mes=TruncMonth("creado"))
        .values("mes")
        .annotate(total=Sum("ganancia"))
        .order_by("mes")
    )

    labels = [i["mes"].strftime("%b %Y") for i in ingresos_por_mes]
    values = [int(i["total"]) for i in ingresos_por_mes]

    mes_max = "Sin datos"
    if ingresos_por_mes:
        mejor = max(ingresos_por_mes, key=lambda x: x["total"])
        mes_max = mejor["mes"].strftime("%B %Y")

    
    subtotal_expr = ExpressionWrapper(
        F("precio_unitario") * F("cantidad"),
        output_field=DecimalField(max_digits=12, decimal_places=2)
    )

    ventas_por_producto = (
        ItemPedido.objects.values("nombre_producto")
        .annotate(
            total_unidades=Sum("cantidad"),
            total_ingresos=Sum(subtotal_expr),
        )
        .order_by("-total_ingresos")
    )

    context = {
        "pedidos": pedidos,
        "total_ganancia": total_ganancia,
        "total_iva": total_iva,
        "cantidad_pedidos": cantidad_pedidos,
        "labels": labels,
        "values": values,
        "mes_max": mes_max,
        "estados": Pedido.Estado.choices,
        "ventas_por_producto": ventas_por_producto,
    }

    return render(request, "estadisticas.html", context)


def custom_404(request, exception):
    return render(request, 'taller/notfound.html')

def _get_or_create_cart(request):
    # Asegura que exista una sesión
    if not request.session.session_key:
        request.session.create()

    if request.user.is_authenticated:
        cart, _ = Carrito.objects.get_or_create(usuario=request.user, activo=True)
        ses_cart = Carrito.objects.filter(clave_sesion=request.session.session_key, activo=True).first()
        if ses_cart and ses_cart != cart:
            for it in ses_cart.items.select_related('producto'):
                tgt, _ = ItemCarrito.objects.get_or_create(
                    carrito=cart, producto=it.producto,
                    defaults={'cantidad': 0, 'precio_unitario': it.precio_unitario}
                )
                tgt.cantidad += it.cantidad
                tgt.precio_unitario = it.precio_unitario
                tgt.save()
            ses_cart.activo = False
            ses_cart.save()
        return cart
    cart, _ = Carrito.objects.get_or_create(clave_sesion=request.session.session_key, activo=True)
    return cart


@require_POST
def api_cart_add(request):
    import json
    data = json.loads(request.body.decode('utf-8'))
    product_id = int(data.get('product_id'))
    quantity = max(1, int(data.get('quantity', 1)))

    p = get_object_or_404(Producto, pk=product_id, activo=True)
    cart = _get_or_create_cart(request)

    item, _ = ItemCarrito.objects.get_or_create(
        carrito=cart, producto=p,
        defaults={'cantidad': 0, 'precio_unitario': p.precio_con_descuento}
    )

    # Respeta stock
    if p.stock is not None:
        new_qty = min(item.cantidad + quantity, p.stock)
    else:
        new_qty = item.cantidad + quantity

    if new_qty == item.cantidad:
        return JsonResponse({"ok": False, "error": "Sin stock"}, status=400)

    item.cantidad = new_qty
    # Congela precio al momento
    item.precio_unitario = p.precio_con_descuento
    item.save()
    track(request, "add_to_cart", product_id=p.id, qty=quantity)
    count = cart.items.aggregate(total=Sum('cantidad'))['total'] or 0
    return JsonResponse({"ok": True, "count": count})


def api_cart_count(request):
    cart = _get_or_create_cart(request)
    count = cart.items.aggregate(total=Sum('cantidad'))['total'] or 0
    return JsonResponse({"count": count})


def api_cart_detail(request):
    cart = _get_or_create_cart(request)
    items = []
    for it in cart.items.select_related('producto'):
        items.append({
            "id": it.producto.id,
            "name": it.producto.nombre,
            "price": float(it.precio_unitario),
            "quantity": it.cantidad,
            "stock": it.producto.stock,
            "image": it.producto.imagen.url if it.producto.imagen else None,
        })
    total = sum(i["price"] * i["quantity"] for i in items)
    return JsonResponse({"items": items, "total": total})


@receiver(user_logged_in)
def _merge_on_login(sender, user, request, **kwargs):
    _get_or_create_cart(request)


def recuperar_contrasena(request):
    return render(request, 'taller/recuperar_contrasena.html')

def recuperar_contra_listo(request):
    return render(request, 'taller/recuperar_contra_listo.html')

def contra_cambiada_exitosa(request):
    return render(request, 'taller/contra_cambiada_exitosa.html')

from django.contrib.auth.views import LoginView
from .forms import EmailAuthenticationForm

class CustomLoginView(LoginView):
    template_name = "taller/login.html"
    authentication_form = EmailAuthenticationForm

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from .models import VehiculoEnVenta
from .forms import VehiculoForm



from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .forms import VehiculoEnVentaForm
from .models import VehiculoImagen


@login_required
def publicar_vehiculo(request):
    if request.method == 'POST':
        form = VehiculoEnVentaForm(request.POST, request.FILES)
        if form.is_valid():
            vehiculo = form.save(commit=False)
            vehiculo.usuario = request.user
            vehiculo.save()

            # Guardar imágenes múltiples
            imagenes = request.FILES.getlist('imagenes')

            if imagenes:
                orden = 0
                for img in imagenes[:10]:  # seguridad extra
                    VehiculoImagen.objects.create(
                        vehiculo=vehiculo,
                        imagen=img,
                        orden=orden
                    )
                    orden += 1

                # Primera imagen como portada
                vehiculo.imagen = imagenes[0]
                vehiculo.save(update_fields=['imagen'])

            messages.success(request, "Tu vehículo fue enviado para aprobación del administrador.")
            return redirect('estado_revi_vehiculos')
        else:
            messages.error(request, "Por favor revisa los campos del formulario.")
    else:
        form = VehiculoEnVentaForm()

    return render(request, 'taller/publicar_vehiculo.html', {'form': form})




@login_required
def estado_revi_vehiculos(request):
    vehiculos = VehiculoEnVenta.objects.filter(usuario=request.user)
    return render(request, 'taller/estado_revi_vehiculos.html', {'vehiculos': vehiculos})



@staff_member_required
def revisar_vehiculo(request):
    pendientes = VehiculoEnVenta.objects.filter(estado='pendiente')
    aprobados = VehiculoEnVenta.objects.filter(estado='aprobado')
    vendidos = VehiculoEnVenta.objects.filter(estado='vendido')
    ocultos = VehiculoEnVenta.objects.filter(estado='oculto')
    return render(request, 'taller/revisar_vehiculo.html', {
        'pendientes': pendientes,
        'aprobados': aprobados,
        'vendidos': vendidos,
        'ocultos': ocultos
    })



@staff_member_required
def aprobar_vehiculo(request, id):
    vehiculo = get_object_or_404(VehiculoEnVenta, id=id)
    vehiculo.estado = 'aprobado'
    vehiculo.save()
    messages.success(request, f"Vehículo {vehiculo.marca} {vehiculo.modelo} aprobado correctamente.")
    return redirect('revisar_vehiculo')


@staff_member_required
def rechazar_vehiculo(request, id):
    vehiculo = get_object_or_404(VehiculoEnVenta, id=id)
    vehiculo.estado = 'rechazado'
    vehiculo.save()
    messages.error(request, f"Vehículo {vehiculo.marca} {vehiculo.modelo} fue rechazado.")
    return redirect('revisar_vehiculo')


@staff_member_required
def cambiar_estado_vehiculo(request, vehiculo_id):
    """Vista AJAX para cambiar el estado de un vehículo"""
    if request.method == 'POST':
        import json
        try:
            vehiculo = get_object_or_404(VehiculoEnVenta, id=vehiculo_id)
            data = json.loads(request.body)
            nuevo_estado = data.get('estado')
            
            estados_validos = ['pendiente', 'aprobado', 'rechazado', 'vendido', 'oculto']
            if nuevo_estado not in estados_validos:
                return JsonResponse({'success': False, 'error': 'Estado no válido'})
            
            vehiculo.estado = nuevo_estado
            vehiculo.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Estado cambiado a {nuevo_estado}'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import Carrito, ItemCarrito

@login_required
def carrito_json(request):
    try:
        # Busca el carrito del usuario activo
        carrito = Carrito.objects.filter(usuario=request.user, activo=True).prefetch_related('items__producto').first()

        if not carrito:
            return JsonResponse({"items": [], "total": 0})

        items_data = []
        total = 0

        for item in carrito.items.all():
            subtotal = float(item.precio_unitario) * item.cantidad
            total += subtotal
            items_data.append({
                "item_id": item.id,
                "producto_id": item.producto.id,
                "nombre": item.producto.nombre,
                "precio": float(item.precio_unitario),
                "cantidad": item.cantidad,
                "subtotal": subtotal
            })

        return JsonResponse({"items": items_data, "total": total})

    except Exception as e:
        print("Error en carrito_json:", e)
        return JsonResponse({"items": [], "total": 0})

from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404
from django.http import JsonResponse

from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required

@require_POST
@login_required
def carrito_actualizar(request, item_id):
    item = get_object_or_404(
        ItemCarrito,
        id=item_id,
        carrito__usuario=request.user,
        carrito__activo=True
    )

   
    try:
        nueva_cantidad = int(request.POST.get("cantidad", 1))
    except (ValueError, TypeError):
        nueva_cantidad = 1

   
    if nueva_cantidad <= 0:
        item.delete()
        return JsonResponse({"ok": True, "msg": "Producto eliminado"})

   
    stock = getattr(item.producto, "stock", None)
    if stock and nueva_cantidad > stock:
        nueva_cantidad = stock

    
    item.cantidad = nueva_cantidad
    item.save(update_fields=["cantidad"])

   
    return JsonResponse({
        "ok": True,
        "msg": "Cantidad actualizada",
        "cantidad": item.cantidad,
        "subtotal": float(item.cantidad * item.precio_unitario)
    })



@require_POST
@login_required
def carrito_eliminar(request, item_id):
    item = get_object_or_404(ItemCarrito, id=item_id, carrito__usuario=request.user, carrito__activo=True)
    item.delete()
    return JsonResponse({"ok": True, "msg": "Producto eliminado"})


# ============================= Bloquear usuario =============================

@staff_member_required
def toggle_bloqueo_usuario(request, user_id):
    usuario = get_object_or_404(Usuario, id=user_id)

    # Evitar bloquear administradores o superusuarios
    if usuario.is_superuser or usuario.rol == Usuario.Rol.ADMIN:
        messages.warning(request, "No puedes bloquear a un usuario administrador.")
        return redirect('admin_usuarios')

    usuario.bloqueado = not usuario.bloqueado
    usuario.save()

    if usuario.bloqueado:
        messages.warning(request, f"El usuario {usuario.get_full_name() or usuario.email} ha sido bloqueado.")
    else:
        messages.success(request, f"El usuario {usuario.get_full_name() or usuario.email} ha sido desbloqueado.")

    return redirect('admin_usuarios')

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import Pedido

@repartidor_or_admin_required
def entregas_view(request):
    """
    Muestra los pedidos asignados al usuario.
    Si el usuario es administrador, muestra todos los pedidos.
    Si es repartidor, muestra solo sus entregas asignadas.
    Excluye pedidos pendientes y cancelados.
    """
    estados_validos = [
        Pedido.Estado.PAGADO,
        Pedido.Estado.PREPARACION,
        Pedido.Estado.EN_RUTA,
        Pedido.Estado.ENVIADO,
        Pedido.Estado.ENTREGADO,
    ]
    
    if request.user.rol == 'ADMIN':
        pedidos = Pedido.objects.filter(
            estado__in=estados_validos
        ).select_related(
            'usuario',
            'direccion_envio',
            'direccion_facturacion',
            'pago'
        ).prefetch_related(
            'items__producto'
        ).order_by('-creado')
    else:
        pedidos = Pedido.objects.filter(
            asignado_a=request.user,
            estado__in=estados_validos
        ).select_related(
            'usuario',
            'direccion_envio',
            'direccion_facturacion',
            'pago'
        ).prefetch_related(
            'items__producto'
        ).order_by('-creado')

    return render(request, 'taller/admin_entregas.html', {'pedidos': pedidos})

from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from django.contrib.admin.views.decorators import staff_member_required
from .models import Pedido

@repartidor_or_admin_required
def actualizar_estado_pedido(request, pedido_id, nuevo_estado):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    # Validar que el usuario tenga permiso para actualizar este pedido
    if request.user.rol == 'REPARTIDOR' and pedido.asignado_a != request.user:
        messages.error(request, "No tienes permiso para actualizar este pedido.")
        return redirect(request.META.get('HTTP_REFERER', 'admin_entregas'))

    if nuevo_estado == "ENTREGADO":
        pedido.marcar_como_entregado(request.user.get_full_name() or request.user.email)
        messages.success(request, f"✅ Pedido entregado por {pedido.entregado_por}.")
    elif nuevo_estado == "EN_RUTA":
        pedido.marcar_en_ruta()
        messages.info(request, "🚚 Pedido marcado como EN RUTA.")
    else:
        pedido.estado = nuevo_estado
        pedido.save()
        messages.info(request, f"📦 Pedido actualizado a {pedido.get_estado_display()}.")

    # Redirige dinámicamente a la página desde donde se hizo clic
    return redirect(request.META.get('HTTP_REFERER', 'admin_entregas'))

import pandas as pd
from django.http import HttpResponse
from django.db.models import Sum
from django.utils import timezone
from django.contrib.admin.views.decorators import staff_member_required
from .models import Pedido
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from .models import Pedido

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
from django.contrib.admin.views.decorators import staff_member_required
from .models import Pedido


@staff_member_required
def descargar_excel_pedidos(request):
    pedidos = Pedido.objects.all().order_by("-creado")

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    # ----------------------------
    # ENCABEZADOS
    # ----------------------------
    headers = [
        "ID", "Cliente", "Estado",
        "Precio (Venta con IVA)",
        "Costo Real",
        "Ganancia Real",
        "IVA Correcto (19%)",
        "Método Entrega", "Método Pago",
        "Fecha", "Entregado por",
        "Productos"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border_style = Side(style="thin", color="D1D5DB")

    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)

    # ----------------------------
    # TOTALES GENERALES
    # ----------------------------
    total_precio = 0
    total_costo = 0
    total_ganancia = 0
    total_iva = 0

    # ----------------------------
    # FILAS
    # ----------------------------
    for p in pedidos:

        # Precio total (con IVA incluido)
        precio = float(p.total or 0)

        # Costo real desde BD
        costo = float(p.costo_total or 0)

        # IVA real (si el precio incluye IVA)
        neto = precio / 1.19
        iva = round(precio - neto)

        # Ganancia real
        ganancia = round(neto - costo)

        # Productos
        items = p.items.all()
        productos_txt = ", ".join([f"{i.nombre_producto} (x{i.cantidad})" for i in items])

        # Totales acumulados
        total_precio += precio
        total_costo += costo
        total_ganancia += ganancia
        total_iva += iva

        ws.append([
            p.id,
            p.usuario.get_full_name() or p.usuario.email,
            p.get_estado_display(),
            precio,
            costo,
            ganancia,
            iva,
            p.get_metodo_entrega_display(),
            p.get_metodo_pago_display(),
            p.creado.strftime("%d/%m/%Y %H:%M"),
            p.entregado_por or "-",
            productos_txt
        ])

    # Formatos de moneda
    for col_letter in ["D", "E", "F", "G"]:
        for cell in ws[col_letter][1:]:
            cell.number_format = '"$"#,##0'

    # Ajustar ancho columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                length = len(str(cell.value))
                max_length = max(max_length, length)
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 3

    # ----------------------------
    # HOJA DE RESUMEN
    # ----------------------------
    resumen = wb.create_sheet(title="Resumen")
    resumen_headers = ["Concepto", "Valor"]
    resumen.append(resumen_headers)

    resumen_datos = [
        ("Total Ventas (Precio con IVA)", total_precio),
        ("Total Costos", total_costo),
        ("Ganancia Total", total_ganancia),
        ("IVA Total", total_iva),
        ("Cantidad de Pedidos", len(pedidos)),
        ("Fecha de Generación", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]

    for item, valor in resumen_datos:
        resumen.append([item, valor])

    # Estilo
    for cell in resumen["A"] + resumen["B"]:
        cell.border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in resumen.columns:
        resumen.column_dimensions[col[0].column_letter].width = 25

    # Encabezado resumen
    for cell in resumen[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Formato moneda
    for cell in resumen["B"][1:5]:
        cell.number_format = '"$"#,##0'

    fecha_str = datetime.now().strftime("%d-%m-%Y")
    nombre_archivo = f"pedidos_financieros_{fecha_str}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'

    wb.save(response)
    return response



from decimal import Decimal
from django.db.models import Sum, F, ExpressionWrapper, DecimalField
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from .models import Pedido

@staff_member_required
def estadisticas_view(request):  
    pedidos = Pedido.objects.prefetch_related("items__producto")


    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    estado = request.GET.get("estado")

    if fecha_inicio:
        pedidos = pedidos.filter(creado__date__gte=fecha_inicio)
    if fecha_fin:
        pedidos = pedidos.filter(creado__date__lte=fecha_fin)
    if estado:
        pedidos = pedidos.filter(estado=estado)


    for p in pedidos:
        costo_total = Decimal(0)
        for i in p.items.all():
            precio_costo = i.producto.costo if i.producto.costo else Decimal(0)
            costo_total += precio_costo * Decimal(i.cantidad or 0)

        precio = Decimal(p.total or 0)
        ganancia = precio - costo_total
        iva = precio * Decimal("0.19")

        p.costo_total = round(costo_total, 2)
        p.ganancia = round(ganancia, 2)
        from math import ceil
        p.iva = Decimal(ceil(iva))

    total_ganancia = sum([p.ganancia for p in pedidos], Decimal(0))
    total_iva = sum([p.iva for p in pedidos], Decimal(0))
    cantidad_pedidos = pedidos.count()

    data_por_mes = (
        pedidos.values("creado__month")
        .annotate(
            ganancia_mes=Sum(
                ExpressionWrapper(F("total") - F("descuento") + F("envio"), output_field=DecimalField())
            )
        )
        .order_by("creado__month")
    )

    labels = [f"{m['creado__month']:02d}" for m in data_por_mes]
    values = [float(m["ganancia_mes"] or 0) for m in data_por_mes]
    mes_max = labels[values.index(max(values))] if values else "Sin datos"

    estados = Pedido.Estado.choices

    context = {
        "pedidos": pedidos,
        "labels": labels,
        "values": values,
        "total_ganancia": total_ganancia,
        "total_iva": total_iva,
        "cantidad_pedidos": cantidad_pedidos,
        "mes_max": mes_max,
        "estados": estados,
    }
    return render(request, "taller/estadisticas.html", context)

from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import CitaForm
from .models import Cita

@login_required
def agendar_servicio(request):
    if request.method == "POST":
        form = CitaForm(request.POST, user=request.user)

        if form.is_valid():
            cita = form.save(commit=False)
            cita.usuario = request.user
            cita.save()

            messages.success(request, "✅ Tu cita fue agendada correctamente.")
            return redirect("mis_citas")
    else:
        form = CitaForm(user=request.user)

    return render(request, "agendar.html", {"form": form})



from django.utils import timezone
from datetime import datetime, timedelta
from .models import BloqueHorario

def generar_bloques_si_faltan():
    """Crea bloques horarios futuros (09:00–18:00) para hoy y mañana si no existen."""
    if not BloqueHorario.objects.exists():
        hoy = timezone.localdate()
        for dia in [hoy, hoy + timedelta(days=1)]:
            for hora in range(9, 19):  # 9:00 a 18:00
                inicio_naive = datetime.combine(dia, datetime.min.time()) + timedelta(hours=hora)
                inicio = timezone.make_aware(inicio_naive, timezone.get_current_timezone())
                fin = inicio + timedelta(hours=1)
                BloqueHorario.objects.get_or_create(inicio=inicio, fin=fin, bloqueado=False)
        print("✅ Bloques horarios creados con zona horaria local.")

def agendar_servicio(request):
    """Vista para agendar citas."""
    from .forms import CitaForm
    from django.contrib import messages

    # 👇 Genera bloques si no hay ninguno
    generar_bloques_si_faltan()

    if request.method == "POST":
        form = CitaForm(request.POST, user=request.user)
        if form.is_valid():
            cita = form.save(commit=False)
            cita.usuario = request.user
            cita.save()
            messages.success(request, "✅ Tu cita fue agendada correctamente. Te contactaremos pronto.")
            return redirect("mis_citas")
    else:
        form = CitaForm(user=request.user)

    return render(request, "agendar.html", {"form": form})


# =============================
#   ELIMINAR PRODUCTO
# =============================
@login_required
def eliminar_producto(request, pk):
    """Eliminar un producto (solo administradores)."""
    if request.user.rol != 'ADMIN':
        messages.error(request, "No tienes permisos para eliminar productos.")
        return redirect('home')
    
    producto = get_object_or_404(Producto, pk=pk)
    nombre_producto = producto.nombre
    producto.delete()
    messages.success(request, f'Producto "{nombre_producto}" eliminado correctamente.')
    return redirect('agregar_editar')


# =============================
#   CÓDIGOS DE DESCUENTO
# =============================
@login_required
def gestionar_codigos_descuento(request):
    """Vista para gestionar códigos de descuento (solo administradores)."""
    if request.user.rol != 'ADMIN':
        messages.error(request, "No tienes permisos para acceder a esta página.")
        return redirect('home')
    
    from .models import CodigoDescuento
    codigos = CodigoDescuento.objects.all().order_by('-creado')
    
    return render(request, 'taller/admin_codigos_descuento.html', {
        'codigos': codigos
    })


@login_required
@require_POST
def aplicar_codigo_descuento(request):
    """Aplicar código de descuento al carrito."""
    from .models import CodigoDescuento, Pedido
    
    codigo = request.POST.get('codigo', '').strip().upper()
    if not codigo:
        return JsonResponse({"ok": False, "msg": "Ingresa un código de descuento."}, status=400)
    
    try:
        codigo_obj = CodigoDescuento.objects.get(codigo=codigo)
    except CodigoDescuento.DoesNotExist:
        return JsonResponse({"ok": False, "msg": "Código de descuento no válido."}, status=400)
    
    valido, mensaje = codigo_obj.es_valido()
    if not valido:
        return JsonResponse({"ok": False, "msg": mensaje}, status=400)
    
    # Obtener carrito
    cart = _get_active_cart(request.user)
    subtotal = cart.subtotal()
    
    # Verificar monto mínimo
    if subtotal < codigo_obj.monto_minimo:
        return JsonResponse({
            "ok": False,
            "msg": f"El monto mínimo para usar este código es ${codigo_obj.monto_minimo:,.0f}"
        }, status=400)
    
    # Calcular descuento
    descuento = codigo_obj.calcular_descuento(subtotal)

    # Verificar límite por usuario (contabiliza usos con compras efectivas o en curso post-pago)
    if codigo_obj.usos_por_usuario is not None:
        estados_validos = [
            Pedido.Estado.PAGADO,
            Pedido.Estado.PREPARACION,
            Pedido.Estado.EN_RUTA,
            Pedido.Estado.ENVIADO,
            Pedido.Estado.ENTREGADO,
        ]
        usos_usuario = Pedido.objects.filter(
            usuario=request.user,
            codigo_descuento=codigo_obj,
            estado__in=estados_validos,
        ).count()
        if usos_usuario >= codigo_obj.usos_por_usuario:
            return JsonResponse({
                "ok": False,
                "msg": f"Ya has utilizado este código el máximo de {codigo_obj.usos_por_usuario} veces."
            }, status=400)
    
    # Guardar en sesión
    request.session['codigo_descuento'] = codigo
    request.session['descuento_aplicado'] = str(descuento)
    
    return JsonResponse({
        "ok": True,
        "msg": f"Código aplicado correctamente. Descuento: ${descuento:,.0f}",
        "descuento": float(descuento),
        "nuevo_total": float(subtotal - descuento)
    })


@login_required
@require_POST
def crear_codigo_descuento(request):
    """Crear un nuevo código de descuento (solo administradores)."""
    if request.user.rol != 'ADMIN':
        return JsonResponse({"ok": False, "msg": "No tienes permisos."}, status=403)
    
    from .models import CodigoDescuento
    
    try:
        codigo = CodigoDescuento.objects.create(
            codigo=request.POST.get('codigo', '').strip().upper(),
            tipo=request.POST.get('tipo', 'PORCENTAJE'),
            valor=Decimal(request.POST.get('valor', 0)),
            monto_minimo=Decimal(request.POST.get('monto_minimo', 0)),
            usos_maximos=int(request.POST.get('usos_maximos')) if request.POST.get('usos_maximos') else None,
            usos_por_usuario=int(request.POST.get('usos_por_usuario')) if request.POST.get('usos_por_usuario') else None,
            inicio=request.POST.get('inicio') or None,
            fin=request.POST.get('fin') or None,
        )
        return JsonResponse({"ok": True, "msg": f"Código {codigo.codigo} creado correctamente."})
    except Exception as e:
        return JsonResponse({"ok": False, "msg": str(e)}, status=400)


# =============================
#   CONTACTO
# =============================
def contacto(request):
    """Página de contacto con formulario."""
    from .models import Contacto
    
    if request.method == 'POST':
        contacto = Contacto.objects.create(
            nombre=request.POST.get('nombre'),
            email=request.POST.get('email'),
            telefono=request.POST.get('telefono', ''),
            asunto=request.POST.get('asunto'),
            mensaje=request.POST.get('mensaje')
        )
        messages.success(request, '¡Gracias por contactarnos! Te responderemos pronto.')
        
        # Opcional: enviar correo al administrador
        # send_mail(...)
        
        return redirect('contacto')
    
    return render(request, 'taller/contacto.html')


@login_required
def admin_contactos(request):
    """Vista para administradores para ver mensajes de contacto."""
    if request.user.rol != 'ADMIN':
        messages.error(request, "No tienes permisos para acceder a esta página.")
        return redirect('home')
    
    from .models import Contacto
    contactos = Contacto.objects.all().order_by('-creado')
    
    return render(request, 'taller/admin_contactos.html', {
        'contactos': contactos
    })


@login_required
@require_POST
def marcar_contacto_leido(request, contacto_id):
    """Marcar un mensaje de contacto como leído."""
    if request.user.rol != 'ADMIN':
        return JsonResponse({"ok": False, "msg": "No tienes permisos."}, status=403)
    
    from .models import Contacto
    contacto = get_object_or_404(Contacto, id=contacto_id)
    contacto.leido = True
    contacto.save()
    
    return JsonResponse({"ok": True})


# =============================
#   RESEÑAS
# =============================
@login_required
@require_POST
def agregar_resena(request, producto_id):
    """Agregar una reseña a un producto."""
    from .models import Resena
    
    producto = get_object_or_404(Producto, id=producto_id)
    
    # Verificar que el usuario haya comprado el producto
    ha_comprado = ItemPedido.objects.filter(
        pedido__usuario=request.user,
        pedido__estado='ENTREGADO',
        producto=producto
    ).exists()
    
    if not ha_comprado:
        return JsonResponse({
            "ok": False,
            "msg": "Solo puedes reseñar productos que hayas comprado."
        }, status=400)
    
    # Verificar que no haya reseñado ya
    if Resena.objects.filter(usuario=request.user, producto=producto).exists():
        return JsonResponse({
            "ok": False,
            "msg": "Ya has reseñado este producto."
        }, status=400)
    
    calificacion = int(request.POST.get('calificacion', 5))
    comentario = request.POST.get('comentario', '')
    
    Resena.objects.create(
        usuario=request.user,
        producto=producto,
        calificacion=calificacion,
        comentario=comentario,
        aprobada=False  # Requiere aprobación del admin
    )
    
    messages.success(request, 'Gracias por tu reseña. Será publicada tras su aprobación.')
    return redirect('producto_detalle', pk=producto_id)


def resenas(request):
    """Página pública de reseñas aprobadas."""
    from .models import Resena, Testimonio

    # Si envían una reseña pública desde el formulario
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        email = request.POST.get('email', '').strip()
        mensaje = request.POST.get('mensaje', '').strip()
        try:
            calificacion = int(request.POST.get('calificacion'))
        except Exception:
            calificacion = None

        if not mensaje:
            messages.error(request, 'El mensaje es obligatorio.')
            return redirect('resenas')

        Testimonio.objects.create(
            usuario=request.user if request.user.is_authenticated else None,
            nombre=nombre,
            email=email,
            mensaje=mensaje,
            calificacion=calificacion,
            aprobada=False
        )

        messages.success(request, 'Gracias por tu reseña. Será publicada tras su aprobación.')
        return redirect('resenas')

    resenas_aprobadas = Resena.objects.filter(aprobada=True).select_related('usuario', 'producto').order_by('-creado')
    resenas_publicas = Testimonio.objects.filter(aprobada=True).select_related('usuario').order_by('-creado')

    return render(request, 'taller/resenas.html', {
        'resenas': resenas_aprobadas,
        'resenas_publicas': resenas_publicas
    })


@login_required
def admin_resenas(request):
    """Vista de administración de reseñas."""
    if request.user.rol != 'ADMIN':
        messages.error(request, "No tienes permisos para acceder a esta página.")
        return redirect('home')
    
    from .models import Resena, Testimonio

    # Reseñas de productos: Pendientes (no aprobadas)
    pendientes = Resena.objects.filter(aprobada=False).select_related('usuario', 'producto').order_by('-creado')
    # Reseñas de productos: Aprobadas
    aprobadas_productos = Resena.objects.filter(aprobada=True).select_related('usuario', 'producto').order_by('-creado')

    # Reseñas públicas: Pendientes (no aprobadas)
    pendientes_resenas = Testimonio.objects.filter(aprobada=False).select_related('usuario').order_by('-creado')
    # Reseñas públicas: Aprobadas/Publicadas
    aprobadas_resenas = Testimonio.objects.filter(aprobada=True).select_related('usuario').order_by('-creado')
    # Reseñas públicas: Rechazadas (None)
    rechazadas_publicas = Testimonio.objects.filter(aprobada=None).select_related('usuario').order_by('-creado')

    # Contadores para badges
    pendientes_count = pendientes.count() + pendientes_resenas.count()
    rechazadas_count = rechazadas_publicas.count()
    aprobadas_count = aprobadas_productos.count() + aprobadas_resenas.count()

    return render(request, 'taller/admin_resenas.html', {
        'pendientes': pendientes,
        'aprobadas_productos': aprobadas_productos,
        'pendientes_resenas': pendientes_resenas,
        'aprobadas_resenas': aprobadas_resenas,
        'rechazadas_publicas': rechazadas_publicas,
        'pendientes_count': pendientes_count,
        'rechazadas_count': rechazadas_count,
        'aprobadas_count': aprobadas_count,
    })



@login_required
@require_POST
def aprobar_testimonio(request, testimonio_id):
    """Aprobar un testimonio (admin)."""
    if request.user.rol != 'ADMIN':
        return JsonResponse({"ok": False, "msg": "No tienes permisos."}, status=403)

    from .models import Testimonio
    test = get_object_or_404(Testimonio, id=testimonio_id)
    test.aprobada = True
    test.save()

    return JsonResponse({"ok": True, "msg": "Reseña aprobada."})


@login_required
@require_POST
def rechazar_testimonio(request, testimonio_id):
    """Rechazar un testimonio público (marcar como rechazado con None)."""
    if request.user.rol != 'ADMIN':
        return JsonResponse({"ok": False, "msg": "No tienes permisos."}, status=403)

    from .models import Testimonio
    test = get_object_or_404(Testimonio, id=testimonio_id)
    test.aprobada = None
    test.save()

    return JsonResponse({"ok": True, "msg": "Reseña rechazada."})


@login_required
@require_POST
@login_required
@require_POST
def eliminar_testimonio(request, testimonio_id):
    """Cambiar estado de testimonio (ocultar si está publicado, publicar si está rechazado)."""
    if request.user.rol != 'ADMIN':
        return JsonResponse({"ok": False, "msg": "No tienes permisos."}, status=403)

    from .models import Testimonio
    test = get_object_or_404(Testimonio, id=testimonio_id)
    # Si está publicado (True), ocultarlo (None). Si está rechazado (None), publicarlo (True)
    test.aprobada = True if test.aprobada is None else None
    test.save()

    return JsonResponse({"ok": True, "msg": "Reseña actualizada."})
@login_required
@require_POST
def aprobar_resena(request, resena_id):
    """Aprobar una reseña."""
    if request.user.rol != 'ADMIN':
        return JsonResponse({"ok": False, "msg": "No tienes permisos."}, status=403)
    
    from .models import Resena
    resena = get_object_or_404(Resena, id=resena_id)
    resena.aprobada = True
    resena.save()
    
    return JsonResponse({"ok": True, "msg": "Reseña aprobada."})


@login_required
@require_POST
def rechazar_resena(request, resena_id):
    """Rechazar una reseña de producto (cambiar estado a no aprobada)."""
    if request.user.rol != 'ADMIN':
        return JsonResponse({"ok": False, "msg": "No tienes permisos."}, status=403)
    
    from .models import Resena
    resena = get_object_or_404(Resena, id=resena_id)
    # Cambiar según el estado actual
    resena.aprobada = not resena.aprobada
    resena.save()
    
    return JsonResponse({"ok": True, "msg": "Reseña actualizada."})


# =============================
#   GALERÍA NOSOTROS
# =============================
@login_required
def admin_galeria_nosotros(request):
    """Gestionar galería de fotos de la página Nosotros."""
    if request.user.rol != 'ADMIN':
        messages.error(request, "No tienes permisos para acceder a esta página.")
        return redirect('home')
    
    from .models import FotoNosotros
    
    if request.method == 'POST' and request.FILES.get('imagen'):
        FotoNosotros.objects.create(
            titulo=request.POST.get('titulo', 'Sin título'),
            descripcion=request.POST.get('descripcion', ''),
            imagen=request.FILES['imagen'],
            orden=int(request.POST.get('orden', 0))
        )
        messages.success(request, 'Foto agregada correctamente.')
        return redirect('admin_galeria_nosotros')
    
    fotos = FotoNosotros.objects.all().order_by('orden', '-creado')
    
    return render(request, 'taller/admin_galeria_nosotros.html', {
        'fotos': fotos
    })


@login_required
@require_POST
def eliminar_foto_nosotros(request, foto_id):
    """Eliminar una foto de la galería."""
    if request.user.rol != 'ADMIN':
        return JsonResponse({"ok": False, "msg": "No tienes permisos."}, status=403)
    
    from .models import FotoNosotros
    foto = get_object_or_404(FotoNosotros, id=foto_id)
    foto.delete()
    
    return JsonResponse({"ok": True, "msg": "Foto eliminada."})


# =============================
#   VEHÍCULOS Y HISTORIAL
# =============================
@login_required
def mis_vehiculos(request):
    """Vista para que el cliente vea sus vehículos y su historial."""
    from .models import VehiculoCliente
    
    vehiculos = VehiculoCliente.objects.filter(usuario=request.user).prefetch_related('historial')
    
    return render(request, 'taller/mis_vehiculos.html', {
        'vehiculos': vehiculos
    })


@login_required
def agregar_vehiculo(request):
    """Agregar un nuevo vehículo."""
    from .models import VehiculoCliente
    
    if request.method == 'POST':
        VehiculoCliente.objects.create(
            usuario=request.user,
            patente=request.POST.get('patente').upper(),
            marca=request.POST.get('marca'),
            modelo=request.POST.get('modelo'),
            año=int(request.POST.get('año')),
            color=request.POST.get('color', ''),
            kilometraje_actual=int(request.POST.get('kilometraje', 0))
        )
        messages.success(request, 'Vehículo agregado correctamente.')
        return redirect('mis_vehiculos')
    
    return render(request, 'taller/agregar_vehiculo.html')


@login_required
def admin_historial_servicios(request):
    """Vista de administración para ver todos los historiales de servicio."""
    if request.user.rol not in ['ADMIN', 'MECANICO']:
        messages.error(request, "No tienes permisos para acceder a esta página.")
        return redirect('home')
    
    from .models import HistorialServicio
    
    historiales = HistorialServicio.objects.all().select_related(
        'vehiculo', 'vehiculo__usuario', 'mecanico_asignado'
    ).order_by('-fecha_ingreso')
    
    return render(request, 'taller/admin_historial_servicios.html', {
        'historiales': historiales
    })


@login_required
@require_POST
def actualizar_historial_servicio(request, historial_id):
    """Actualizar el estado de un servicio (para mecánicos)."""
    if request.user.rol not in ['ADMIN', 'MECANICO']:
        return JsonResponse({"ok": False, "msg": "No tienes permisos."}, status=403)
    
    from .models import HistorialServicio
    
    historial = get_object_or_404(HistorialServicio, id=historial_id)
    
    nuevo_estado = request.POST.get('estado')
    comentario = request.POST.get('comentario_mecanico', '')
    
    if nuevo_estado:
        historial.estado = nuevo_estado
    if comentario:
        historial.comentario_mecanico = comentario
    if nuevo_estado == 'completado' and not historial.fecha_salida:
        historial.fecha_salida = timezone.now()
    
    historial.save()
    
    messages.success(request, 'Historial actualizado correctamente.')
    return redirect('admin_historial_servicios')


# =============================
#   SWITCH VISTA ADMIN/CLIENTE
# =============================
@login_required
def toggle_vista_admin(request):
    """Cambiar entre vista de administrador y vista de cliente."""
    if request.user.rol != 'ADMIN':
        messages.error(request, "No tienes permisos para cambiar de vista.")
        return redirect('home')
    
    # Usar sesión para guardar la preferencia
    vista_actual = request.session.get('vista_modo', 'cliente')
    nueva_vista = 'admin' if vista_actual == 'cliente' else 'cliente'
    request.session['vista_modo'] = nueva_vista
    
    messages.success(request, f'Cambiado a vista de {nueva_vista}.')
    return redirect('home')


# ==================== MARKETPLACE DE VEHÍCULOS ====================

def vehiculos_venta(request):
    """Vista pública para ver todos los vehículos en venta aprobados"""
    vehiculos = VehiculoEnVenta.objects.filter(estado='aprobado').order_by('-fecha_publicacion')
    
    # Filtros
    marca = request.GET.get('marca', '')
    anio_desde = request.GET.get('anio_desde', '')
    precio_max = request.GET.get('precio_max', '')
    combustible = request.GET.get('combustible', '')
    
    if marca:
        vehiculos = vehiculos.filter(marca__icontains=marca)
    if anio_desde:
        try:
            vehiculos = vehiculos.filter(**{"año__gte": int(anio_desde)})
        except ValueError:
            pass
    if precio_max:
        try:
            vehiculos = vehiculos.filter(precio__lte=int(precio_max))
        except ValueError:
            pass
    if combustible:
        # Mapear alias comunes a los valores del modelo
        alias = {
            'gasolina': 'bencina',
            'bencina': 'bencina',
            'diesel': 'diesel',
            'diésel': 'diesel',
        }
        comb_val = alias.get(combustible.lower())
        if comb_val:
            vehiculos = vehiculos.filter(combustible=comb_val)
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(vehiculos, 9)  # 9 vehículos por página
    page = request.GET.get('page')
    vehiculos = paginator.get_page(page)
    
    return render(request, 'taller/vehiculos_venta.html', {
        'vehiculos': vehiculos
    })


def vehiculo_detalle(request, vehiculo_id):
    """Vista de detalle de un vehículo en venta"""
    vehiculo = get_object_or_404(VehiculoEnVenta, id=vehiculo_id, estado='aprobado')
    
    # Cargar galería de imágenes adicionales
    imagenes_galeria = vehiculo.imagenes.all().order_by('orden', 'fecha_subida')
    
    return render(request, 'taller/vehiculo_detalle.html', {
        'vehiculo': vehiculo,
        'imagenes_galeria': imagenes_galeria,
    })

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
import base64

from .models import Pedido
from .forms import ConfirmarEntregaForm

@repartidor_or_admin_required
def confirmar_entrega(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)

    # Validar que el usuario tenga permiso para confirmar esta entrega
    if request.user.rol == 'REPARTIDOR' and pedido.asignado_a != request.user:
        messages.error(request, "No tienes permiso para confirmar esta entrega.")
        return redirect('admin_entregas')
 
    if pedido.metodo_entrega == Pedido.MetodoEntrega.RETIRO:
        pedido.estado = Pedido.Estado.ENTREGADO
        pedido.entregado_por = request.user.get_full_name() or request.user.username
        pedido.hora_entrega = timezone.localtime().time()
        pedido.save()
        messages.success(request, "✅ Pedido marcado como entregado (retiro en tienda)")
        return redirect("admin_entregas")


    if request.method == "POST":
        form = ConfirmarEntregaForm(request.POST, request.FILES, instance=pedido)

        # Capturar firma en base64
        firma_base64 = request.POST.get("firma_entrega_hidden")

        if form.is_valid():
            pedido = form.save(commit=False)
            pedido.estado = Pedido.Estado.ENTREGADO
            pedido.entregado_por = request.user.get_full_name() or request.user.username
            pedido.hora_entrega = timezone.localtime().time()

            # Convertir base64 → archivo si existe
            if firma_base64:
                format, imgstr = firma_base64.split(';base64,')
                ext = format.split('/')[-1]
                pedido.firma_entrega = ContentFile(base64.b64decode(imgstr), name=f"firma_{pedido.id}.{ext}")

            pedido.save()

            messages.success(request, "✅ Entrega registrada con evidencia")
            return redirect("admin_entregas")

        messages.error(request, "❌ Revisa los datos del formulario")

    return redirect("admin_entregas")

from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect
from .models import HorarioDia
from .forms import HorarioDiaForm  # Lo crearemos abajo

@staff_member_required
def admin_horarios(request):
    from .models import HorarioDia, HoraDisponible
    from django.utils import timezone
    
    dias_semana = {
        0: "Lunes",
        1: "Martes",
        2: "Miércoles",
        3: "Jueves",
        4: "Viernes",
        5: "Sábado",
        6: "Domingo",
    }

    dias = {}
    for numero, nombre in dias_semana.items():
        horarios = HorarioDia.objects.filter(dia_semana=numero, activo=True).order_by("hora_inicio")
        
        # Para cada horario, obtener las horas disponibles registradas
        horas_por_horario = {}
        for h in horarios:
            horas_disponibles = h.horas_disponibles.filter(disponible=True).order_by("-fecha", "hora")
            horas_por_horario[h.id] = horas_disponibles
        
        dias[numero] = {
            "nombre": nombre,
            "horarios": horarios,
            "horas_por_horario": horas_por_horario
        }

    return render(request, "taller/admin_horarios.html", {
        "dias": dias,
    })


@login_required
def api_generar_bloques_por_fecha(request):
    """Genera (si hace falta) y devuelve los bloques disponibles para una fecha.
    Retorna JSON: { ok: True, bloques: [{id, hora}, ...] }
    """
    from .models import HoraDisponible, BloqueHorario, Cita
    from datetime import datetime as _dt, datetime

    fecha_str = request.GET.get("fecha")
    if not fecha_str:
        return JsonResponse({"ok": False, "error": "Falta parámetro fecha"}, status=400)

    try:
        fecha = _dt.strptime(fecha_str, "%Y-%m-%d").date()
    except Exception:
        return JsonResponse({"ok": False, "error": "Formato de fecha inválido (YYYY-MM-DD)"}, status=400)

    try:
        tz = timezone.get_current_timezone()
        horas = HoraDisponible.objects.filter(fecha=fecha, disponible=True)

        bloques_list = []
        for hd in horas:
            inicio_naive = datetime.combine(fecha, hd.hora)
            inicio = timezone.make_aware(inicio_naive, tz)
            fin = inicio + timedelta(hours=1)

            bloque, _ = BloqueHorario.objects.get_or_create(inicio=inicio, fin=fin, defaults={"bloqueado": False})

            # Solo devolver bloques libres
            if not bloque.bloqueado and not Cita.objects.filter(bloque=bloque, estado=Cita.Estado.RESERVADA).exists():
                bloques_list.append({"id": bloque.id, "hora": inicio.strftime("%H:%M")})

        # Ordenar por hora
        bloques_list = sorted(bloques_list, key=lambda x: x["hora"])

        return JsonResponse({"ok": True, "bloques": bloques_list})
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=500)


# =============================
# ENDPOINTS AJAX PARA HORAS DISPONIBLES
# =============================

@staff_member_required
@require_POST
def crear_hora_disponible(request):
    """Crea una hora disponible para agendamiento."""
    from .models import HorarioDia, HoraDisponible
    from datetime import datetime
    
    try:
        horario_id = request.POST.get('horario_id')
        fecha_str = request.POST.get('fecha')
        hora_str = request.POST.get('hora')
        
        if not all([horario_id, fecha_str, hora_str]):
            return JsonResponse({'success': False, 'error': 'Faltan datos'}, status=400)
        
        # Validar que el horario existe y está activo
        horario = HorarioDia.objects.get(id=horario_id, activo=True)
        
        # Parsear fecha y hora
        from datetime import datetime, time
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        hora = datetime.strptime(hora_str, '%H:%M').time()
        
        # Validar que la hora está dentro del rango del horario del día
        if not (horario.hora_inicio <= hora < horario.hora_fin):
            return JsonResponse({
                'success': False, 
                'error': f'La hora debe estar entre {horario.hora_inicio} y {horario.hora_fin}'
            }, status=400)
        
        # Crear o actualizar hora disponible
        hora_disp, created = HoraDisponible.objects.get_or_create(
            fecha=fecha,
            hora=hora,
            defaults={'horario_dia': horario, 'disponible': True}
        )
        
        if not created:
            # Si ya existe, marcar como disponible
            hora_disp.disponible = True
            hora_disp.save(update_fields=['disponible'])
            mensaje = "Hora actualizada como disponible"
        else:
            mensaje = "Hora registrada como disponible"
        
        return JsonResponse({
            'success': True,
            'message': mensaje,
            'hora_id': hora_disp.id,
            'fecha': fecha_str,
            'hora': hora_str
        })
    
    except HorarioDia.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Horario no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@staff_member_required
@require_POST
def eliminar_hora_disponible(request):
    """Marca una hora como no disponible (ocupada)."""
    from .models import HoraDisponible
    
    try:
        hora_id = request.POST.get('hora_id')
        
        if not hora_id:
            return JsonResponse({'success': False, 'error': 'Falta hora_id'}, status=400)
        
        hora = HoraDisponible.objects.get(id=hora_id)
        hora.disponible = False
        hora.save(update_fields=['disponible'])
        
        return JsonResponse({
            'success': True,
            'message': 'Hora marcada como ocupada'
        })
    
    except HoraDisponible.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Hora no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@staff_member_required
def listar_horas_disponibles(request):
    """Lista horas disponibles por horario (AJAX)."""
    from .models import HorarioDia, HoraDisponible
    
    try:
        horario_id = request.GET.get('horario_id')
        fecha = request.GET.get('fecha')
        
        if not all([horario_id, fecha]):
            return JsonResponse({'success': False, 'error': 'Faltan parámetros'}, status=400)
        
        # Validar horario existe
        HorarioDia.objects.get(id=horario_id)
        
        # Obtener horas disponibles
        horas = HoraDisponible.objects.filter(
            horario_dia_id=horario_id,
            fecha=fecha,
            disponible=True
        ).values('id', 'hora').order_by('hora')
        
        return JsonResponse({
            'success': True,
            'horas': list(horas)
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
